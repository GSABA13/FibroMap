"""
Lecture du fichier Excel contenant les données de prélèvements amiante.

Ce module fournit la fonction `charger_excel` qui ouvre un classeur Excel
(.xlsx) en mode lecture seule, lit la feuille "Prv Am" et retourne la liste
des prélèvements sous forme d'objets Echantillon.

Correspondance des colonnes Excel (feuille "Prv Am") :
  A  – Unités
  B  – Zone
  C  – Équipements
  D  – Localisation
  E  – Élément sondé
  F  – Description
  G  – Prélèvement       ← clé : ligne ignorée si vide
  H  – Date
  I  – Résultat
  J  – (non utilisé)
  K  – (non utilisé)
  L  – Volume
  M  – Photo
  N  – Étage
  O  – Référence Plan
  P  – Marquage
  Q  – N° Prv Labo
  R  – Commentaires

Les deux premières lignes (en-têtes) sont ignorées.
Si la feuille "Prv Am" est absente, la fonction retourne une liste vide
sans lever d'exception.
"""

import logging

import openpyxl

from src.models.echantillon import Echantillon
from src.models.planche import Planche
from src.services.couleur_resolver import resoudre_couleur
from src.services.legende_builder import construire_texte

# Journalisation du module
logger = logging.getLogger(__name__)

# Nom de la feuille cible dans le classeur Excel
NOM_FEUILLE: str = "Prv Am"

# Indices de colonnes fixes (base 1, correspondant à l'API openpyxl)
COL_LOCALISATION: int = 4    # D
COL_ELEMENT_SONDE: int = 5   # E
COL_DESCRIPTION: int = 6     # F
COL_PRELEVEMENT: int = 7     # G
COL_RESULTAT: int = 9        # I
COL_REFERENCE_PLAN: int = 15  # O

# Numéro de la première ligne de données (après les en-têtes)
# Les 4 premières lignes sont réservées aux en-têtes du fichier Excel
PREMIERE_LIGNE_DONNEES: int = 5

# Libellé de la colonne clé primaire tel qu'il apparaît en ligne 4 de l'Excel
# (peut être en colonne K ou M selon le type de fichier — détection automatique)
LIBELLE_ID_PRIMAIRE: str = "Identifiant photo"


def _detecter_colonne_id(feuille, libelle: str) -> int | None:
    """
    Cherche dans la ligne d'en-tête (PREMIERE_LIGNE_DONNEES - 1) la colonne
    dont la valeur correspond au libellé donné.

    Retourne l'indice en base 1 de la colonne trouvée, ou None si absent.
    Cela permet de supporter différents formats d'Excel où la clé primaire
    peut se trouver en colonne K ou M selon le type de fichier.
    """
    ligne_entete = PREMIERE_LIGNE_DONNEES - 1
    for row in feuille.iter_rows(min_row=ligne_entete, max_row=ligne_entete):
        for cellule in row:
            if cellule.value and str(cellule.value).strip() == libelle:
                logger.debug(
                    "Colonne '%s' détectée en colonne %d.", libelle, cellule.column
                )
                return cellule.column
    logger.warning("Colonne '%s' introuvable dans les en-têtes — id_primaire sera vide.", libelle)
    return None


def _valeur_cellule(row: tuple, index_1base: int) -> str:
    """
    Extrait la valeur d'une cellule dans une ligne openpyxl et la convertit
    en chaîne de caractères propre.

    Args:
        row         : tuple de cellules retourné par iter_rows().
        index_1base : numéro de colonne en base 1 (1 = colonne A).

    Returns:
        La valeur sous forme de str, ou une chaîne vide si la cellule est None.
    """
    cellule = row[index_1base - 1]  # Conversion base-1 → base-0
    valeur = cellule.value
    if valeur is None:
        return ""
    return str(valeur).strip()


def charger_excel(chemin: str) -> list[Echantillon]:
    """
    Charge et analyse le fichier Excel des prélèvements amiante.

    Ouvre le classeur en mode read_only pour minimiser la consommation mémoire,
    lit toutes les lignes valides de la feuille "Prv Am" (colonne G non vide,
    à partir de la ligne 3) et retourne la liste des Echantillon construits.

    Args:
        chemin: Chemin absolu ou relatif vers le fichier Excel (.xlsx).

    Returns:
        Liste d'objets Echantillon, dans l'ordre de lecture du fichier.
        Retourne une liste vide si :
          - la feuille "Prv Am" est absente du classeur,
          - aucune ligne valide n'est trouvée.

    Raises:
        FileNotFoundError : si le fichier Excel n'existe pas à l'emplacement indiqué.
        openpyxl.utils.exceptions.InvalidFileException : si le fichier n'est pas
            un classeur Excel valide.
    """
    logger.info("Chargement du fichier Excel : %s", chemin)

    echantillons: list[Echantillon] = []

    # openpyxl 3.x ne supporte pas le context manager sur load_workbook
    # → on utilise try/finally pour garantir la fermeture dans tous les cas
    # keep_vba=True permet d'ouvrir les fichiers .xlsm (Excel avec macros) sans erreur
    classeur = openpyxl.load_workbook(chemin, read_only=True, data_only=True, keep_vba=True)
    try:
        # Vérification de la présence de la feuille cible
        if NOM_FEUILLE not in classeur.sheetnames:
            logger.warning(
                "Feuille '%s' absente du classeur '%s'. Retour d'une liste vide.",
                NOM_FEUILLE,
                chemin,
            )
            return []

        feuille = classeur[NOM_FEUILLE]

        # Détection automatique de la colonne clé primaire par son libellé en ligne 4
        col_id_primaire: int | None = _detecter_colonne_id(feuille, LIBELLE_ID_PRIMAIRE)

        # Parcours des lignes à partir de la première ligne de données
        for numero_ligne, row in enumerate(
            feuille.iter_rows(min_row=PREMIERE_LIGNE_DONNEES), start=PREMIERE_LIGNE_DONNEES
        ):
            # Extraction de l'identifiant du prélèvement (colonne G)
            prelevement: str = _valeur_cellule(row, COL_PRELEVEMENT)

            # On ignore les lignes dont la colonne G est vide
            if not prelevement:
                logger.debug("Ligne %d ignorée : colonne G vide.", numero_ligne)
                continue

            # Extraction des autres champs utiles
            description: str = _valeur_cellule(row, COL_DESCRIPTION)
            resultat: str = _valeur_cellule(row, COL_RESULTAT)
            localisation: str = _valeur_cellule(row, COL_LOCALISATION)
            element_sonde: str = _valeur_cellule(row, COL_ELEMENT_SONDE)
            reference_plan: str = _valeur_cellule(row, COL_REFERENCE_PLAN)
            id_primaire: str = _valeur_cellule(row, col_id_primaire) if col_id_primaire else ""

            # Résolution de la couleur et de la mention selon le résultat
            couleur, mention = resoudre_couleur(resultat)

            # Construction des trois lignes de texte de la bulle de légende
            texte_ligne1, texte_ligne2, texte_ligne3 = construire_texte(
                prelevement=prelevement,
                description=description,
                resultat=resultat,
                localisation=localisation,
                element_sonde=element_sonde,
            )

            # Instanciation de l'objet Echantillon
            echantillon = Echantillon(
                prelevement=prelevement,
                description=description,
                resultat=resultat,
                localisation=localisation,
                element_sonde=element_sonde,
                reference_plan=reference_plan,
                couleur=couleur,
                mention=mention,
                texte_ligne1=texte_ligne1,
                texte_ligne2=texte_ligne2,
                texte_ligne3=texte_ligne3,
                id_primaire=id_primaire,
            )

            echantillons.append(echantillon)
            logger.debug(
                "Ligne %d – prélèvement '%s' chargé (résultat : '%s', mention : '%s').",
                numero_ligne,
                prelevement,
                resultat,
                mention,
            )
    finally:
        classeur.close()

    logger.info(
        "%d prélèvement(s) chargé(s) depuis '%s'.", len(echantillons), chemin
    )

    # Log de diagnostic — vérification du chargement
    logger.info("[DIAGNOSTIC] %d échantillon(s) chargé(s) depuis '%s'", len(echantillons), chemin)
    if echantillons:
        premier = echantillons[0]
        logger.debug(
            "[DIAGNOSTIC] Premier échantillon — prélèvement='%s', localisation='%s', "
            "résultat='%s', mention='%s', couleur=%s",
            premier.prelevement,
            premier.localisation,
            premier.resultat,
            premier.mention,
            premier.couleur,
        )
    else:
        logger.warning("[DIAGNOSTIC] Liste d'échantillons VIDE après lecture de '%s'", chemin)

    return echantillons


def maj_bulles_depuis_echantillons(
    planches: list[Planche],
    echantillons: list[Echantillon],
) -> int:
    """
    Met à jour les bulles de légende de toutes les planches avec les nouvelles
    données Excel, en faisant correspondre sur l'identifiant de prélèvement.

    Pour chaque bulle dont l'échantillon existe dans la nouvelle liste,
    les données (textes, couleur, mention…) sont remplacées par les valeurs
    fraîches. La position et le point d'ancrage de la bulle sont conservés.

    Si un prélèvement n'existe plus dans le nouvel Excel, la bulle est conservée
    telle quelle et un avertissement est journalisé.

    Paramètres
    ----------
    planches      : liste des planches du chantier (toutes, pas seulement l'active)
    echantillons  : liste des échantillons issus du rechargement Excel

    Retourne
    --------
    Nombre de bulles mises à jour.
    """
    index = {e.id_primaire: e for e in echantillons if e.id_primaire}
    nb = 0
    for planche in planches:
        for bulle in planche.bulles:
            if bulle.echantillon is None:
                continue
            cle = bulle.echantillon.id_primaire
            nouvel_ech = index.get(cle)
            if nouvel_ech is not None:
                bulle.echantillon = nouvel_ech
                bulle.couleur_rgb = nouvel_ech.couleur
                nb += 1
            else:
                logger.warning(
                    "Prélèvement '%s' introuvable dans le nouvel Excel "
                    "— bulle de la planche '%s' conservée.",
                    bulle.echantillon.prelevement,
                    planche.reference_plan,
                )
    return nb
